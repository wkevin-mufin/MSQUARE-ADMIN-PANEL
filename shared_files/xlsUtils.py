import openpyxl


def getRowCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def getColumnCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def readDAta(filePath, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value


def writeData(filePath, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(filePath)
